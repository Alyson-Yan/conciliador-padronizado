import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# COLUNAS PADRÃO DO RELATÓRIO FINAL
# ============================================================

COLUNAS_RELATORIO_PADRAO = [
    "autorizacao_erp",
    "nsu_erp",
    "valor_bruto_erp",
    "valor_liquido_erp",
    "chave_erp",
    "data_emissao_erp",
    "cliente_erp",

    "instituicao",
    "valor_bruto_instituicao",
    "valor_liquido_instituicao",
    "data_venda_instituicao",
    "data_vencimento_instituicao",
    "data_pagamento_instituicao",
    "autorizacao_instituicao",
    "nsu_instituicao",
    "parcela_instituicao",
    "total_parcelas_instituicao",

    "diferenca_dias",
    "diferenca_valor",
    "status_conciliacao",
    "pontuacao",
]


# ============================================================
# SEPARAÇÃO E RESUMO
# ============================================================

def separar_conciliados(df_conciliado):
    df_conciliado = df_conciliado.copy()

    mascara_conciliado = df_conciliado["status_conciliacao"].str.startswith(
        "Conciliado",
        na=False
    )

    df_aba_conciliados = df_conciliado[mascara_conciliado].copy()
    df_aba_nao_conciliados = df_conciliado[~mascara_conciliado].copy()

    # Casos especiais não devem ficar na aba "Não conciliados".
    # Eles vão para abas próprias.
    if "tipo_lancamento_instituicao" in df_aba_nao_conciliados.columns:
        tipo_lancamento = (
            df_aba_nao_conciliados["tipo_lancamento_instituicao"]
            .astype(str)
            .str.lower()
        )

        mascara_especial = (
            tipo_lancamento.str.contains("aluguel", na=False)
            | tipo_lancamento.str.contains("tarifa", na=False)
            | tipo_lancamento.str.contains("estorno", na=False)
            | tipo_lancamento.str.contains("cancelamento", na=False)
            | tipo_lancamento.str.contains("chargeback", na=False)
        )

        df_aba_nao_conciliados = df_aba_nao_conciliados[
            ~mascara_especial
        ].copy()

    return df_aba_conciliados, df_aba_nao_conciliados


def calcular_totais(df):
    return {
        "valor_liquido": df["valor_liquido_instituicao"].sum(),
        "valor_bruto": df["valor_bruto_instituicao"].sum(),
        "quantidade": len(df),
    }


def criar_resumo(df_conciliados, df_nao_conciliados, abas_especiais=None):
    if not isinstance(abas_especiais, dict):
        abas_especiais = {}

    totais_conciliados = calcular_totais_dataframe(df_conciliados)
    totais_nao_conciliados = calcular_totais_dataframe(df_nao_conciliados)

    totais_especiais = {}

    for nome_aba, df_especial in abas_especiais.items():
        if isinstance(df_especial, pd.DataFrame):
            totais_especiais[nome_aba] = calcular_totais_dataframe(df_especial)

    valor_liquido_total_geral = (
        totais_conciliados["valor_liquido"]
        + totais_nao_conciliados["valor_liquido"]
        + sum(total["valor_liquido"] for total in totais_especiais.values())
    )

    valor_bruto_total_geral = (
        totais_conciliados["valor_bruto"]
        + totais_nao_conciliados["valor_bruto"]
        + sum(total["valor_bruto"] for total in totais_especiais.values())
    )

    quantidade_total_geral = (
        totais_conciliados["quantidade"]
        + totais_nao_conciliados["quantidade"]
        + sum(total["quantidade"] for total in totais_especiais.values())
    )

    linhas_resumo = [
        ["RELATORIO DE CONCILIACAO", "", ""],
        ["", "", ""],

        ["TOTAL GERAL", "", ""],
        ["Valor liquido total geral", "", valor_liquido_total_geral],
        ["Valor bruto total geral", "", valor_bruto_total_geral],
        ["Quantidade total de titulos", "", quantidade_total_geral],
        ["", "", ""],

        ["CONCILIADOS", "", ""],
        ["Valor liquido total", "", totais_conciliados["valor_liquido"]],
        ["Valor bruto total", "", totais_conciliados["valor_bruto"]],
        ["Quantidade de titulos", "", totais_conciliados["quantidade"]],
        ["", "", ""],

        ["NAO CONCILIADOS", "", ""],
        ["Valor liquido total", "", totais_nao_conciliados["valor_liquido"]],
        ["Valor bruto total", "", totais_nao_conciliados["valor_bruto"]],
        ["Quantidade de titulos", "", totais_nao_conciliados["quantidade"]],
        ["", "", ""],
    ]

    for nome_aba, totais in totais_especiais.items():
        linhas_resumo.extend([
            [nome_aba.upper(), "", ""],
            ["Valor liquido total", "", totais["valor_liquido"]],
            ["Valor bruto total", "", totais["valor_bruto"]],
            ["Quantidade de titulos", "", totais["quantidade"]],
            ["", "", ""],
        ])

    return pd.DataFrame(
        linhas_resumo,
        columns=["categoria", "descricao", "valor"]
    )

    
def calcular_totais_dataframe(df):
    total_liquido = 0
    total_bruto = 0
    quantidade = 0

    if df is None:
        return {
            "valor_liquido": 0,
            "valor_bruto": 0,
            "quantidade": 0,
        }
        

    # Se vier uma Series por engano, transforma em DataFrame com 1 linha.
    if isinstance(df, pd.Series):
        df = df.to_frame().T

    if not isinstance(df, pd.DataFrame):
        return {
            "valor_liquido": 0,
            "valor_bruto": 0,
            "quantidade": 0,
        }

    if df.empty:
        return {
            "valor_liquido": 0,
            "valor_bruto": 0,
            "quantidade": 0,
        }

    if "valor_liquido_instituicao" in df.columns:
        total_liquido = pd.to_numeric(
            df["valor_liquido_instituicao"],
            errors="coerce"
        ).fillna(0).sum()

    if "valor_bruto_instituicao" in df.columns:
        total_bruto = pd.to_numeric(
            df["valor_bruto_instituicao"],
            errors="coerce"
        ).fillna(0).sum()

    quantidade = len(df)

    return {
        "valor_liquido": total_liquido,
        "valor_bruto": total_bruto,
        "quantidade": quantidade,
    }
# ============================================================
# ABAS ESPECIAIS
# ============================================================

def criar_abas_especiais(df_conciliado):
    abas_especiais = {}

    if "tipo_lancamento_instituicao" not in df_conciliado.columns:
        return abas_especiais

    tipo_lancamento = (
        df_conciliado["tipo_lancamento_instituicao"]
        .astype(str)
        .str.lower()
    )

    df_cancelamentos = df_conciliado[
        tipo_lancamento.str.contains("cancelamento", na=False)
        | tipo_lancamento.str.contains("chargeback", na=False)
    ].copy()

    if not df_cancelamentos.empty:
        abas_especiais["Cancelamentos"] = df_cancelamentos

    df_aluguel = df_conciliado[
        tipo_lancamento.str.contains("aluguel", na=False)
        | tipo_lancamento.str.contains("tarifa", na=False)
    ].copy()

    if not df_aluguel.empty:
        abas_especiais["Aluguel e Tarifas"] = df_aluguel

    df_estornos = df_conciliado[
        tipo_lancamento.str.contains("estorno", na=False)
    ].copy()

    if not df_estornos.empty:
        abas_especiais["Estornos"] = df_estornos

    return abas_especiais


# ============================================================
# COLUNAS AMIGÁVEIS DO RELATÓRIO
# ============================================================

def nome_coluna_relatorio(coluna, instituicao):
    nomes_fixos = {
        "autorizacao_erp": "autorizacao erp",
        "nsu_erp": "codigo nsu erp",
        "valor_bruto_erp": "valor bruto erp",
        "valor_liquido_erp": "valor liquido erp",
        "chave_erp": "chave erp",
        "data_emissao_erp": "data transacao erp",
        "cliente_erp": "pessoa do titulo",

        "instituicao": "instituicao financeira",

        "diferenca_dias": "dif dias",
        "diferenca_valor": "dif valor",
        "status_conciliacao": "status",
        "pontuacao": "pontuacao",
    }

    nomes_instituicao = {
        "valor_bruto_instituicao": f"valor bruto {instituicao}",
        "valor_liquido_instituicao": f"valor liquido {instituicao}",
        "data_venda_instituicao": f"data venda {instituicao}",
        "data_vencimento_instituicao": f"data vencimento {instituicao}",
        "data_pagamento_instituicao": f"data recebimento {instituicao}",
        "autorizacao_instituicao": f"autorizacao {instituicao}",
        "nsu_instituicao": f"nsu {instituicao}",
        "parcela_instituicao": f"parcela atual {instituicao}",
        "total_parcelas_instituicao": f"total parcelas {instituicao}",
    }

    if coluna in nomes_fixos:
        return nomes_fixos[coluna]

    if coluna in nomes_instituicao:
        return nomes_instituicao[coluna]

    return (
        coluna
        .replace("_instituicao", f" {instituicao}")
        .replace("_", " ")
    )

def preparar_colunas_relatorio(df, instituicao):
    df = df.copy()

    for coluna in COLUNAS_RELATORIO_PADRAO:
        if coluna not in df.columns:
            df[coluna] = pd.NA

    df = df[COLUNAS_RELATORIO_PADRAO]

    novos_nomes = {
        coluna: nome_coluna_relatorio(coluna, instituicao)
        for coluna in COLUNAS_RELATORIO_PADRAO
    }

    df = df.rename(columns=novos_nomes)

    return df


# ============================================================
# CHAVES ERP NO RESUMO
# ============================================================

def inserir_chaves_erp_em_blocos(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)

    if "Conciliados" not in wb.sheetnames or "Resumo" not in wb.sheetnames:
        wb.save(caminho_arquivo)
        return

    ws_conciliados = wb["Conciliados"]
    ws_resumo = wb["Resumo"]

    header = [cell.value for cell in ws_conciliados[1]]

    if "chave erp" not in header:
        wb.save(caminho_arquivo)
        return

    indice_chave = header.index("chave erp") + 1

    chaves = []

    for row in range(2, ws_conciliados.max_row + 1):
        valor = ws_conciliados.cell(row=row, column=indice_chave).value

        if valor is not None and str(valor).strip() != "":
            chaves.append(str(valor))

    blocos = [
        chaves[i:i + 2000]
        for i in range(0, len(chaves), 2000)
    ]

    linha_inicial = ws_resumo.max_row + 2

    if blocos:
        ws_resumo.cell(
            row=linha_inicial,
            column=1,
            value="CHAVES ERP CONCILIADAS"
        )
        ws_resumo.cell(row=linha_inicial, column=1).font = Font(bold=True)

    for indice, bloco in enumerate(blocos, start=1):
        linha = linha_inicial + indice

        ws_resumo.cell(row=linha, column=1, value=f"Grupo {indice}")
        ws_resumo.cell(row=linha, column=2, value=", ".join(bloco))

    wb.save(caminho_arquivo)


# ============================================================
# FORMATAÇÃO DO EXCEL
# ============================================================

def aplicar_formatacao_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)

    preenchimento_cabecalho = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    fonte_cabecalho = Font(bold=True)

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]

        for cell in ws[1]:
            cell.font = fonte_cabecalho
            cell.fill = preenchimento_cabecalho
            cell.alignment = Alignment(horizontal="center")

        for coluna in ws.columns:
            maior_tamanho = 0
            letra_coluna = get_column_letter(coluna[0].column)

            for cell in coluna:
                valor = cell.value

                if valor is not None:
                    maior_tamanho = max(maior_tamanho, len(str(valor)))

                if cell.row > 1:
                    nome_cabecalho = str(
                        ws.cell(row=1, column=cell.column).value
                    ).lower()

                    if "data" in nome_cabecalho:
                        cell.number_format = "DD/MM/YYYY"

                    if "valor" in nome_cabecalho or "taxa" in nome_cabecalho:
                        if isinstance(cell.value, (int, float)):
                            texto_categoria = ""

                            if nome_aba == "Resumo":
                                texto_categoria = str(
                                    ws.cell(row=cell.row, column=1).value
                                ).lower()

                            if "quantidade" in texto_categoria:
                                cell.number_format = "0"
                            else:
                                cell.number_format = 'R$ #,##0.00'

            ws.column_dimensions[letra_coluna].width = min(
                maior_tamanho + 2,
                50
            )

    wb.save(caminho_arquivo)


# ============================================================
# FUNÇÃO PRINCIPAL DO RELATÓRIO
# ============================================================

def gerar_relatorio_conciliacao(
    df_conciliado,
    caminho_saida,
    instituicao,
    incluir_abas_especiais=True,
    incluir_chaves_erp=True,
    aplicar_formatacao=True,
):
    df_conciliado = df_conciliado.copy()

    # =========================
    # SEPARAR ABAS PRINCIPAIS
    # =========================

    df_aba_conciliados, df_aba_nao_conciliados = separar_conciliados(
        df_conciliado
    )

    # =========================
    # CRIAR ABAS ESPECIAIS
    # =========================

    if incluir_abas_especiais:
        abas_especiais = criar_abas_especiais(df_conciliado)
    else:
        abas_especiais = {}

    # =========================
    # CRIAR RESUMO COM DADOS ORIGINAIS
    # =========================
    # Importante:
    # O resumo deve usar os DataFrames ANTES de preparar_colunas_relatorio(),
    # porque preparar_colunas_relatorio renomeia as colunas para o usuário final.

    df_resumo = criar_resumo(
        df_aba_conciliados,
        df_aba_nao_conciliados,
        abas_especiais
    )

    # =========================
    # TOTALIZADORES DO RETORNO
    # =========================

    valor_liquido_conciliado = df_aba_conciliados[
        "valor_liquido_instituicao"
    ].sum()

    valor_liquido_nao_conciliado = df_aba_nao_conciliados[
        "valor_liquido_instituicao"
    ].sum()

    qtd_conciliados = len(df_aba_conciliados)
    qtd_nao_conciliados = len(df_aba_nao_conciliados)

    # =========================
    # FORMATAR ABAS PARA USUÁRIO FINAL
    # =========================

    df_aba_conciliados_formatado = preparar_colunas_relatorio(
        df_aba_conciliados,
        instituicao
    )

    df_aba_nao_conciliados_formatado = preparar_colunas_relatorio(
        df_aba_nao_conciliados,
        instituicao
    )

    abas_especiais_formatadas = {}

    for nome_aba, df_aba in abas_especiais.items():
        abas_especiais_formatadas[nome_aba] = preparar_colunas_relatorio(
            df_aba,
            instituicao
        )

    # =========================
    # ESCREVER EXCEL
    # =========================

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_aba_conciliados_formatado.to_excel(
            writer,
            sheet_name="Conciliados",
            index=False
        )

        df_aba_nao_conciliados_formatado.to_excel(
            writer,
            sheet_name="Não conciliados",
            index=False
        )

        for nome_aba, df_aba_formatada in abas_especiais_formatadas.items():
            df_aba_formatada.to_excel(
                writer,
                sheet_name=nome_aba,
                index=False
            )

        # Resumo sempre como última aba
        df_resumo.to_excel(
            writer,
            sheet_name="Resumo",
            index=False
        )

    if incluir_chaves_erp:
        inserir_chaves_erp_em_blocos(caminho_saida)

    if aplicar_formatacao:
        aplicar_formatacao_excel(caminho_saida)

    return {
        "caminho_saida": caminho_saida,
        "qtd_conciliados": qtd_conciliados,
        "qtd_nao_conciliados": qtd_nao_conciliados,
        "valor_liquido_conciliado": valor_liquido_conciliado,
        "valor_liquido_nao_conciliado": valor_liquido_nao_conciliado,
    }